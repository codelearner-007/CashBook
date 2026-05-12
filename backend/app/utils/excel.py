from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette ────────────────────────────────────────────────────────────────
TEAL        = "39AAAA"
TEAL_DARK   = "2B8080"
TEAL_LIGHT  = "F4FAFA"
TEAL_MID    = "DFF0F0"
GREEN       = "15803D"
GREEN_LIGHT = "DCFCE7"
RED         = "B91C1C"
RED_LIGHT   = "FEE2E2"
GREY        = "F8FAFC"
BORDER_CLR  = "E2E8F0"
DARK        = "0F172A"
MUTED       = "64748B"
WHITE       = "FFFFFF"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _side(style="thin", color=BORDER_CLR) -> Side:
    return Side(border_style=style, color=color)


def _border(style="thin") -> Border:
    s = _side(style)
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_border() -> Border:
    th = _side("medium", TEAL)
    tn = _side("thin")
    return Border(left=tn, right=tn, top=th, bottom=th)


def _cell(ws, row, col, value=None, *, bold=False, size=9, color=DARK,
          fill=None, align="left", valign="center", fmt=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=False)
    if fill:   c.fill   = fill
    if fmt:    c.number_format = fmt
    if border: c.border = border
    return c


_PAYMENT_LABELS = {"cash": "Cash", "online": "Online", "cheque": "Cheque", "other": "Other"}
_TYPE_LABELS    = {"in": "Cash In", "out": "Cash Out"}


def generate_excel(book_name: str, currency: str, entries: list, summary: dict,
                   date_from=None, date_to=None, filters: dict = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    wb.properties.creator = "CashBook"

    sym  = (currency or "").strip()
    cfmt = f'"{sym} " #,##0.00'
    net  = summary["net_balance"]

    # ═══════════════════════════════════════════════════════════════════════
    # ROW 1 — App banner
    # ═══════════════════════════════════════════════════════════════════════
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "CashBook  —  Financial Report"
    c.font  = Font(bold=True, size=16, color=WHITE)
    c.fill  = _fill(TEAL)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    # ═══════════════════════════════════════════════════════════════════════
    # ROW 2 — Book & period info
    # ═══════════════════════════════════════════════════════════════════════
    ws.merge_cells("A2:I2")
    end_date = date_to or datetime.now().strftime("%Y-%m-%d")
    c = ws["A2"]
    c.value = f"  {book_name}  ·  {date_from or 'All time'}  →  {end_date}  ·  {len(entries)} transactions"
    c.font  = Font(size=9, color=WHITE)
    c.fill  = _fill(TEAL_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # ═══════════════════════════════════════════════════════════════════════
    # ROW 3 — Active filters
    # ═══════════════════════════════════════════════════════════════════════
    filter_items = []
    if filters:
        if filters.get("entry_type"):
            filter_items.append(f"Type: {_TYPE_LABELS.get(filters['entry_type'], filters['entry_type'].title())}")
        if filters.get("contact_name"):
            filter_items.append(f"Contact: {filters['contact_name']}")
        if filters.get("category"):
            filter_items.append(f"Category: {filters['category']}")
        if filters.get("payment_mode"):
            filter_items.append(f"Payment: {_PAYMENT_LABELS.get(filters['payment_mode'], filters['payment_mode'].title())}")

    ws.merge_cells("A3:I3")
    fc = ws["A3"]
    if filter_items:
        fc.value = "  Filters: " + "  ·  ".join(filter_items)
        fc.font  = Font(size=8.5, bold=True, color=TEAL_DARK)
    else:
        fc.value = "  Filters: All entries (no additional filters)"
        fc.font  = Font(size=8.5, italic=True, color=MUTED)
    fc.fill      = _fill(TEAL_LIGHT)
    fc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 16

    # ═══════════════════════════════════════════════════════════════════════
    # ROWS 4-5 — Summary block
    # ═══════════════════════════════════════════════════════════════════════
    summary_items = [
        ("TOTAL INCOME",   summary["total_in"],  GREEN, GREEN_LIGHT),
        ("TOTAL EXPENSES", summary["total_out"], RED,   RED_LIGHT),
        ("NET BALANCE",    net, GREEN if net >= 0 else RED,
         GREEN_LIGHT if net >= 0 else RED_LIGHT),
    ]
    for idx, (label, value, val_color, bg) in enumerate(summary_items):
        col = idx + 1

        # Label cell (row 4)
        lc = ws.cell(row=4, column=col, value=label)
        lc.font      = Font(bold=True, size=8, color=WHITE)
        lc.fill      = _fill(TEAL)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border    = _border()

        # Value cell (row 5)
        vc = ws.cell(row=5, column=col, value=value)
        vc.font          = Font(bold=True, size=13, color=val_color)
        vc.fill          = _fill(bg)
        vc.number_format = cfmt
        vc.alignment     = Alignment(horizontal="center", vertical="center")
        vc.border        = _thick_border()

    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 24

    # ═══════════════════════════════════════════════════════════════════════
    # ROW 7 — Column headers
    # ═══════════════════════════════════════════════════════════════════════
    HEADERS = ["Date", "Time", "Remark", "Category", "Contact", "Payment Mode",
               "Cash In", "Cash Out", "Balance"]
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=7, column=col, value=h)
        c.font      = Font(bold=True, color=WHITE, size=9)
        c.fill      = _fill(TEAL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()
    ws.row_dimensions[7].height = 20

    # ═══════════════════════════════════════════════════════════════════════
    # DATA ROWS — from row 8
    # ═══════════════════════════════════════════════════════════════════════
    running   = 0.0
    total_in  = 0.0
    total_out = 0.0

    for row_idx, e in enumerate(entries, 8):
        amt   = float(e["amount"])
        is_in = e["type"] == "in"

        if is_in:
            running  += amt
            total_in += amt
            in_val, out_val = amt, None
        else:
            running   -= amt
            total_out += amt
            in_val, out_val = None, amt

        row_bg = _fill(GREY) if row_idx % 2 == 0 else _fill(WHITE)

        bd = _border("hair")

        def _d(col, val, fmt=None, bold=False, val_color=DARK, center=False):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill      = row_bg
            c.font      = Font(size=9, bold=bold, color=val_color)
            c.alignment = Alignment(horizontal="center" if center else "left",
                                    vertical="center")
            c.border    = bd
            if fmt: c.number_format = fmt
            return c

        _d(1, str(e.get("entry_date", ""))[:10])
        _d(2, str(e.get("entry_time") or "")[:5], center=True)
        _d(3, e.get("remark") or "")
        _d(4, e.get("category") or "")
        _d(5, e.get("contact_name") or "")
        _d(6, e.get("payment_mode") or "")

        if in_val is not None:
            _d(7, in_val, cfmt, bold=True, val_color=GREEN, center=True)
        else:
            _d(7, None)

        if out_val is not None:
            _d(8, out_val, cfmt, bold=True, val_color=RED, center=True)
        else:
            _d(8, None)

        bal_color = GREEN if running >= 0 else RED
        _d(9, running, cfmt, bold=True, val_color=bal_color, center=True)

        ws.row_dimensions[row_idx].height = 16

    # ═══════════════════════════════════════════════════════════════════════
    # TOTALS ROW
    # ═══════════════════════════════════════════════════════════════════════
    tr = len(entries) + 8
    for col in range(1, 10):
        c = ws.cell(row=tr, column=col)
        c.fill   = _fill(TEAL)
        c.font   = Font(bold=True, color=WHITE, size=9)
        c.border = _border()
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=tr, column=1, value="TOTAL")
    ws.cell(row=tr, column=7, value=total_in).number_format  = cfmt
    ws.cell(row=tr, column=8, value=total_out).number_format = cfmt
    ws.cell(row=tr, column=9, value=running).number_format   = cfmt
    ws.row_dimensions[tr].height = 18

    # ═══════════════════════════════════════════════════════════════════════
    # FOOTER NOTE
    # ═══════════════════════════════════════════════════════════════════════
    note_row = tr + 2
    ws.merge_cells(f"A{note_row}:I{note_row}")
    nc = ws.cell(row=note_row, column=1,
                 value=f"Generated by CashBook  ·  {datetime.now().strftime('%B %d, %Y  %I:%M %p')}")
    nc.font      = Font(size=8, color="94A3B8", italic=True)
    nc.alignment = Alignment(horizontal="left", vertical="center")

    # ═══════════════════════════════════════════════════════════════════════
    # COLUMN WIDTHS  /  FREEZE  /  FILTER
    # ═══════════════════════════════════════════════════════════════════════
    widths = [13, 8, 30, 15, 15, 16, 16, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A8"
    if entries:
        ws.auto_filter.ref = f"A7:I{tr - 1}"

    ws.sheet_properties.tabColor = TEAL

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
